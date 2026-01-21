from flask import Flask, request
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)

FILE = "attendance.xlsx"

# ---------- UID TO STUDENT MAPPING ----------
students = {
    "3017b120": (1, "VIJAY ANAND", "EEE"),
    "11e5d51d": (2, "SATHISH", "EEE"),
    "e336f62e": (3, "GURU PRASAD", "EEE"),
    "d3e1112e": (4, "HARI PRASAD", "EEE"),
}

# ---------- Create or Open Excel ----------
if os.path.exists(FILE):
    wb = load_workbook(FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["SNO", "NAME", "DEPARTMENT", "DATE", "TIME"])
    wb.save(FILE)

@app.route('/mark', methods=['GET'])
def mark():
    uid = request.args.get('uid').lower()

    if uid not in students:
        print("Unknown Card:", uid)
        return "UNKNOWN CARD"

    sno, name, dept = students[uid]

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    ws.append([sno, name, dept, date, time])
    wb.save(FILE)

    print(f"Attendance Marked: {sno}, {name}, {dept}, {date}, {time}")
    return "OK"

if __name__ == "__main__":
    print("Attendance Server Running on Port 5000")
    app.run(host="0.0.0.0", port=5000)
